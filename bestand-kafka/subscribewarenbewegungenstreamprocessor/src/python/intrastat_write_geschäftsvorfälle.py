import collections
import json
from openpyxl.reader.excel import load_workbook

class IntrastatWriteGeschäftsvorfälle:
    def __init__(self):
        self.read_excel_data()
        # self.write_json()

    def read_excel_data(self):
        file = "geschäftsvorfälle.xlsx"
        wb = load_workbook(file, data_only=True)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        rows = ws.max_row
        columns = ws.max_column
        row_dict = {}

        headerdata_list = []
        for column in range(1, columns + 1):
            headerdata_list.append(ws.cell(row=1, column=column).value)

        for row in range(2, rows + 1):
            celldata_list = []
            for column in range(1, columns + 1):
                cell_value = ws.cell(row=row, column=column).value
                celldata_list.append(cell_value)

            row_dict = collections.OrderedDict(zip(headerdata_list, celldata_list))
            jsonstr = json.dumps(row_dict).encode("utf8")
            print(jsonstr)

if __name__ == "__main__":
    intrastatWriteGeschäftsvorfälle = IntrastatWriteGeschäftsvorfälle()